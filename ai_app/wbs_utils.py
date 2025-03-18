import tempfile
import time
from decimal import Decimal
import requests
import re
import os
from pathlib import Path
from docx import Document
from decouple import config
import openpyxl
from .common import log_execution_time
from .utils import get_file_content, process_docx_content


@log_execution_time
def get_wbs_content(access_token, item_id):
    wbs_drive_id = config("WBS_DRIVE")
    url = f"https://graph.microsoft.com/v1.0/drives/{wbs_drive_id}/items/{item_id}"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(url, headers=headers)
    response = response.json()
    download_url = response.get("@microsoft.graph.downloadUrl", None)
    if not download_url:
        raise "There was an issue while getting the Download URL from Sharepoint"

    # Download the file
    file_response = requests.get(download_url)
    if file_response.status_code != 200:
        raise Exception("Failed to download the WBS file from SharePoint")

    # Save the file locally
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        temp_filename = temp_file.name
        temp_file.write(file_response.content)

    # Read the Excel content
    wbs_content = read_tasks_from_excel(temp_filename)

    # Cleanup the temporary file
    os.remove(temp_filename)

    return wbs_content


def upload_wbs_to_sharepoint(access_token, file_path, project_id):
    try:
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        site_id = config("SITE_ID")
        wbs_drive_id = config("WBS_DRIVE")
        # Upload the file
        upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{wbs_drive_id}/root:/WBS-{project_id}.xlsx:/content"

        with open(file_path, "rb") as file:
            response = requests.put(upload_url, headers=headers, data=file)

        if response.status_code not in [200, 201]:
            raise Exception(f"Failed to upload file: {response.json()}")

        # Extract the uploaded file's item ID
        item_id = response.json().get("id")

        # Get existing columns
        columns_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{wbs_drive_id}/items/{item_id}/listItem/fields"
        print(columns_url)
        fields_response = requests.get(columns_url, headers=headers)

        if fields_response.status_code != 200:
            raise Exception(f"Failed to fetch columns: {fields_response.json()}")

        updated_project_id = {
            "ProjetID":project_id
        }

        # Update the columns
        update_response = requests.patch(columns_url, headers=headers, json=updated_project_id)

        if update_response.status_code != 200:
            raise Exception(f"Failed to update project_id: {update_response.json()}")

    except Exception as e:
        raise Exception(f"Error during SharePoint upload or update: {str(e)}")


def save_costs_to_existing_excel(costs, file_path):
    """
    Writes Azure service cost data into an existing Excel file.
    If the file does not exist, it creates a new one.
    It adds the data to a sheet called 'Cost Breakdown'.

    :param costs: Dictionary containing 'breakdown' with service names, costs, SKU names, and regions.
    :param file_path: Path to the .xlsx file.
    """
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = openpyxl.Workbook()

    # Check if 'Cost Breakdown' sheet exists, otherwise create it
    if "Cost Breakdown" in wb.sheetnames:
        ws = wb["Cost Breakdown"]
    else:
        ws = wb.create_sheet(title="Cost Breakdown")
        ws.append(["Service Name", "Cost (USD)", "SKU Name", "Region"])  # Add headers if new sheet

    # Append new data from the costs dictionary
    for service_name, cost_data in costs.get("breakdown", {}).items():
        sku_name = cost_data.get("skuName", "N/A")
        region = cost_data.get("region", "N/A")
        cost = cost_data.get("cost", 0)

        ws.append([service_name, cost, sku_name, region])

    # Save and close the workbook properly
    wb.save(file_path)
    wb.close()
    print(f"Data successfully written to {file_path} in 'Cost Breakdown' sheet.")


def save_cost_dict_list_to_excel(data_list, file_path):
    """
    Writes a list of dictionaries into an Excel file.
    If the file does not exist, it creates a new one.
    It adds the data to a sheet called 'Cost Breakdown'.

    :param data_list: List of dictionaries with the same keys.
    :param file_path: Path to the .xlsx file.
    """
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = openpyxl.Workbook()

    # Check if 'Cost Breakdown' sheet exists, otherwise create it
    if "Cost Breakdown" in wb.sheetnames:
        ws = wb["Cost Breakdown"]
    else:
        ws = wb.create_sheet(title="Cost Breakdown")
        # Write headers based on the first dictionary keys
        headers = list(data_list[0].keys())
        ws.append(headers)

    # Append rows from the list of dictionaries
    for row in data_list:
        ws.append([str(row.get(col, "")) if isinstance(row.get(col), Decimal) else row.get(col, "") for col in headers])

    # Save and close the workbook properly
    wb.save(file_path)
    
    wb.close()
    print(f"Data successfully written to {file_path} in 'Cost Breakdown' sheet.")


@log_execution_time
def create_upload_wbs(access_token, result, project_id, costs):
    output_file_path = create_file(result, project_id)
    # save_costs_to_existing_excel(costs, output_file_path)
    save_cost_dict_list_to_excel(costs, output_file_path)

    # Upload to SharePoint
    upload_wbs_to_sharepoint(access_token, output_file_path, project_id)
    # update_current_step(project_id, "Questionnaire Review")

    # Remove the file after successful submission
    os.remove(output_file_path)

    return True


def add_tasks_to_excel(file_path, phases_data, project_id, sheet_name="Eng WBS"):
    """
    Loads an Excel file, navigates to a specific sheet, and adds task data from the phases_data dictionary,
    inserting hours and task titles into specific columns for each phase.

    :param file_path: Path to the Excel file.
    :param phases_data: Dictionary containing phase-wise hours and tasks.
    :param project_id: Project ID for naming the file.
    :param sheet_name: Name of the worksheet to modify.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Select the worksheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        # Define starting row and column mappings for each phase
        phases_columns = {
            "phase1": ["C", "D", 11],
            "phase2": ["H", "I", 9],
            "phase3": ["M", "N", 9],
            "phase4": ["R", "S", 9]
        }

        # Iterate over each phase and add data to corresponding columns
        for phase, (hours_col, task_col, start_row) in phases_columns.items():
            if phase in phases_data:
                hours_list = phases_data[phase].get("hours", [])
                tasks_list = phases_data[phase].get("tasks", [])

                for i, (hours, task) in enumerate(zip(hours_list, tasks_list)):
                    row = start_row + i
                    sheet[f"{hours_col}{row}"] = hours  # Hours estimate
                    sheet[f"{task_col}{row}"] = task  # Task title

        # Save the updated file
        file_name = f"wbs_{project_id}.xlsx"
        time.sleep(4)
        wb.save(file_name)
        print(f"Data successfully added to '{sheet_name}' in {file_name}")
        return file_name

    except Exception as e:
        print(f"Error: {e}")
        return None

def create_file(ai_response, project_id):
    file_path = "ai_app/wbs.xlsx"
    json_ai_response = eval(ai_response)
    file_name = add_tasks_to_excel(file_path, json_ai_response, project_id)
    return file_name


def read_tasks_from_excel(file_path, sheet_name="Eng WBS"):
    """
    Reads tasks and hours from an Excel sheet and returns them in a structured dictionary.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the worksheet to read from.
    :return: Dictionary containing phase-wise tasks and hours.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        # Define starting row and column mappings for each phase
        phases_columns = {
            "phase1": ["C", "D", 11],
            "phase2": ["H", "I", 9],
            "phase3": ["M", "N", 9],
            "phase4": ["R", "S", 9]
        }

        phases_data = {}

        for phase, (hours_col, task_col, start_row) in phases_columns.items():
            hours_list = []
            tasks_list = []
            row = start_row

            while True:
                hours_cell = sheet[f"{hours_col}{row}"].value
                task_cell = sheet[f"{task_col}{row}"].value

                if hours_cell is None and task_cell is None:
                    break  # Stop if both columns are empty

                if hours_cell is not None:
                    hours_list.append(hours_cell)
                else:
                    hours_list.append("")  # Maintain index alignment

                if task_cell is not None:
                    tasks_list.append(task_cell)
                else:
                    tasks_list.append("")

                row += 1

            if hours_list or tasks_list:
                phases_data[phase] = {"hours": hours_list, "tasks": tasks_list}

        wb.close()
        return phases_data

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None
